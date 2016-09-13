# -*- coding: utf-8 -*-
"""
Created on Wed Sep  7 10:19:10 2016

@author: Haiyu

整合全部景区2015/9/1——2016/9/1 包含江苏剔除南京的31省客流量
"""

import os
import datetime
import xlrd
from openpyxl.workbook import Workbook

keys = ["安徽",
"北京",
"福建",
"甘肃",
"广东",
"广西",
"贵州",
"海南",
"河北",
"河南",
"黑龙江",
"湖北",
"湖南",
"吉林",
"江苏",
"江西",
"辽宁",
"内蒙古",
"宁夏",
"青海",
"山东",
"山西",
"陕西",
"上海",
"四川",
"天津",
"西藏",
"新疆",
"云南",
"浙江",
"重庆"];

wb = Workbook()
ws = wb.get_active_sheet()


def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)

    sheet = book.sheet_by_index(0)
    nrows = sheet.nrows
    ncols = sheet.ncols

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()
    if nrows == 0 or ncols == 0:
        return book1
    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)
            #print(sheet.cell_value(row, col))

    return book1
    

def EraseEmpty(str):
    res = ""
    for c in str:
        if c != " ":
            res += c
            
    return res
    
def CombineData(file, cnt, day):
    book = open_xls_as_xlsx(file)
    sheet = book.get_active_sheet()
    
    ws.cell(row=cnt,column=1).value = day
    if sheet.max_column == 1 or sheet.max_row == 1:
        return
    data = {}
    for r in range(2, sheet.max_row + 1):
        data[EraseEmpty(sheet.cell(row=r, column=1).value)] = sheet.cell(row=r, column=2).value
    
    c = 2

    for key in keys:
        ws.cell(row=1,column=c).value = key
        ws.cell(row=cnt,column=c).value = data.get(key)
        c += 1

    
def main():
    day = datetime.date(2016, 4, 1)
    cnt = 2
    dayFrom = day.strftime("%Y-%m-%d")
    dayEnd = datetime.date(2016,9,2).strftime("%Y-%m-%d")

    while (dayFrom != dayEnd):
        print(dayFrom)
        file = os.getcwd() + "\\Data2\\" + dayFrom + "--" + dayFrom + ".xls"
        CombineData(file, cnt, dayFrom)
        day = day + datetime.timedelta(days=1)
        cnt += 1
        dayFrom = day.strftime("%Y-%m-%d")
    wb.save("test.xlsx")
    
    
if __name__ == '__main__':
    main()
