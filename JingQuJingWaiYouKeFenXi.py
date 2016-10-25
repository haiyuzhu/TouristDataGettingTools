# -*- coding: utf-8 -*-

"""
Created on  2016/10/25
Author:     Haiyu Zhu
E-mail:     zhuhaiyu1991@163.com
"""

from Downloader import *
import requests
import os
import ExcelTools
import datetime
from openpyxl.workbook import Workbook


class JingQuJingWaiYouKeFenXi(Downloader):
    def __init__(self, **kwargs):
        Downloader.__init__(self, **kwargs)
        self.scenes = {
            100: "夫子庙",
            124: "南京市博物馆(朝天宫)",
            250129: "大报恩寺",
            250010: "南京博物院",
            110: "钟山风景区",
            121: "总统府",
            123: "玄武湖",
            126: "梅园新村",
            194: "红山动物园",
            127: "栖霞山",
            608: "珍珠泉",
            250004: "求雨山文化园",
            250005: "雨发生态园",
            128: "南京大屠杀遇难同胞纪念馆",
            122: "阅江楼",
            802: "高淳国际慢城",
            250008: "游子山",
            250009: "武家嘴农业科技园",
            133: "高淳老街",
            250003: "金牛湖",
            250001: "巴布洛生态谷",
            250002: "大泉村",
            611: "明文化村",
            251292: "牛首山",
            251293: "银杏湖",
            251294: "黄龙岘",
            251295: "石塘",
            251296: "大塘金",
            251297: "紫清湖",
            120: "雨花台",
            129: "南京科技馆",
            801: "周园",
            250006: "傅家边农业生态科技园",
            250007: "天生桥景区"}

    def fetch_data(self, **kwargs):
        if ("day_head" not in kwargs) or ("day_tail" not in kwargs) or ("scenic_id" not in kwargs) \
                or ("position_name" not in kwargs) or ("file_dir" not in kwargs):
            print("Arguments must contain: day_head, day_tail, scenic_id, position_name!")
            return
        day_head = kwargs["day_head"]
        day_tail = kwargs["day_tail"]
        scenic_id = kwargs["scenic_id"]
        position_name = kwargs["position_name"]
        file_dir = kwargs["file_dir"]

        resp = requests.post(url=self.host + "/kyfx/info/oldHeader.jsp",
                             params={"scenicid": scenic_id, "position_name": position_name},
                             headers={"cookie": self.cookie})
        resp.raise_for_status()

        resp = requests.post(url=self.host + "/kyfx/info/foreigncountry.jsp",
                             params={"var": 4, "data": (day_head + "---" + day_tail)},
                             headers={"cookie": self.cookie})
        resp.raise_for_status()

        resp = requests.post(url=self.host + "/kyfx/info/xiazai.jsp",
                             params={"var": "country"}, headers={"cookie": self.cookie})
        resp.raise_for_status()

        file_name = file_dir + "\\" + position_name + day_head + "--" + day_tail + ".xls"
        file = open(file_name, "wb+")
        file.write(resp.content)
        file.close()

    def process_data(self, **kwargs):
        if ("day_head" not in kwargs) or ("day_tail" not in kwargs):
            print("Arguments must contain: day_head, day_tail!")
            return
        day_head = kwargs["day_head"]
        day_tail = kwargs["day_tail"]
        country_to_idx = {}
        for (scene_id, scene_name) in self.scenes.items():
            print("Processing...", scene_name)
            data_dir = self.work_dir + "\\" + scene_name
            if not os.path.exists(data_dir):
                print("No directory named", scene_name)
                continue
            wb = Workbook()
            ws = wb.active

            cur_day = day_head
            cur_day_str = cur_day.strftime("%Y-%m-%d")
            day_t = day_tail + datetime.timedelta(days=1)
            day_t_str = day_t.strftime("%Y-%m-%d")
            cur_rows = 2

            while cur_day_str != day_t_str:
                file_name = data_dir + "\\" + cur_day + "--" + cur_day + ".xls"
                self.__combine_data(ws,file_name,country_to_idx, cur_rows, cur_day_str)
                cur_rows += 1
                cur_day = cur_day + datetime.timedelta(days=1)
                cur_day_str = cur_day.strftime("%Y-%m-%d")
            file_to_save = data_dir + "\\" + scene_name + day_head.strftime("%Y-%m-%d") + "--" \
                           + day_tail.strftime("%Y-%m-%d") + ".xlsx"
            for (country, idx) in country_to_idx.items():
                ws.cell(row=1, column=idx).value = country
            wb.save(file_to_save)

    def __combine_data(self, ws, file_name, country_to_idx, cur_rows, cur_day_str):
        book = ExcelTools.open_xls_as_xlsx(file_name)
        sheet = book.active
        ws.cell(row=cur_rows, column=1).value = cur_day_str

        if sheet.max_column == 1 or sheet.max_row == 1:
            return

        for r in range(2, sheet.max_row + 1):
            country = ExcelTools.erase_space(sheet.cell(row=r, column=2).value)
            num = sheet.cell(row=r, column=3).value

            idx = country_to_idx.get(country)
            if idx is None:
                idx = len(country_to_idx) + 2
                country_to_idx[country] = idx
            ws.cell(row=cur_rows, column=idx).value = num

    def pipeline(self, continuous=False, set_cookie=True, **kwargs):
        if ("year_head" not in kwargs) or ("year_tail" not in kwargs) or ("month_head" not in kwargs) \
                or ("month_tail" not in kwargs) or ("day_head" not in kwargs) or ("day_tail" not in kwargs):
            print("Arguments must contain: year_head, year_tail, month_head, month_tail, day_head, day_tail")
            return
        if set_cookie:
            self.set_cookie()
        else:
            if not self.login():
                return

        year_head = kwargs["year_head"]
        year_tail = kwargs["year_tail"]
        month_head = kwargs["month_head"]
        month_tail = kwargs["month_tail"]
        day_head = kwargs["day_head"]
        day_tail = kwargs["day_tail"]

        day_h = datetime.date(year_head, month_head, day_head)
        day_t = datetime.date(year_tail, month_tail, day_tail)
        if continuous:
            day_h_str = day_h.strftime("%Y-%m-%d")
            day_t_str = day_t.strftime("%Y-%m-%d")
            for (scene_id, scene_name) in self.scenes.items():
                file_dir = self.work_dir + "\\" + scene_name
                if not os.path.exists(file_dir):
                    os.makedirs(file_dir)
                self.fetch_data(day_head=day_h_str, day_tail=day_t_str, scenic_id=scene_id,
                                position_name=scene_name, file_dir=file_dir)
                print("Downloading...", scene_name, day_h_str, "to", day_t_str )
        else:
            for (scene_id, scene_name) in self.scenes.items():
                file_dir = self.work_dir + "\\" + scene_name
                if not os.path.exists(file_dir):
                    os.makedirs(file_dir)
                cur_day = day_h
                cur_day_str = cur_day.strftime("%Y-%m-%d")
                day_t_str = (day_t + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                while cur_day_str != day_t_str:
                    self.fetch_data(day_head=cur_day_str, day_tail=cur_day_str, scenic_id=scene_id,
                                    position_name=scene_name, file_dir=file_dir)
                    print("Downloading...", scene_name, cur_day_str)
                    cur_day = cur_day + datetime.timedelta(days=1)
                    cur_day_str = cur_day.strftime("%Y-%m-%d")
                self.process_data(day_head=day_h, day_tail=day_t)


if __name__ == "__main__":
    downloader = JingQuJingWaiYouKeFenXi()
    downloader.pipeline(continuous=True, set_cookie=True,year_head=2016, year_tail=2016, month_head=9, month_tail=9,
                        day_head=1, day_tail=4)