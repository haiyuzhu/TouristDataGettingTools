# -*- coding: utf-8 -*-

"""
Created on  2016/10/24
Author:     Haiyu Zhu
E-mail:     zhuhaiyu1991@163.com
"""

import requests
import os
import datetime
import time
import Downloader as Base
import ExcelTools
from openpyxl.workbook import Workbook


# 下载全市景区客源排名里面的数据
class QuanShiJingQuKeYuanPaiMing(Base.Downloader):

    def __init__(self, **kwargs):
        Base.Downloader.__init__(self, **kwargs)
        self.data_path = self.work_dir + "\\" + time.strftime("Month%m-day%d-%H-%M-%S", time.localtime())
        print("Data will be saved in:", self.data_path)
        if not os.path.exists(self.data_path):
            os.makedirs(self.data_path)

        self.include_jiangsu = False
        self.include_nanjing = False
        self.provinces = ["安徽",
            "北京",
            "福建",
            "甘肃",
            "广东",
            "广西",
            "贵州",
            "海南",
            "河北",
            "河南",
            "黑龙江",
            "湖北",
            "湖南",
            "吉林",
            "江苏",
            "江西",
            "辽宁",
            "内蒙古",
            "宁夏",
            "青海",
            "山东",
            "山西",
            "陕西",
            "上海",
            "四川",
            "天津",
            "西藏",
            "新疆",
            "云南",
            "浙江",
            "重庆"]

    def fetch_data(self, **kwargs):
        # include_nanjing, include_jiangsu, day_head, day_tail
        day_head = ""
        day_tail = ""
        if "day_head" not in kwargs:
            print("Arguments must contain day_head!")
            return

        day_head = kwargs["day_head"]
        if "day_tail" in kwargs:
            day_tail = kwargs["day_tail"]
        else:
            print("waring: Downloading data from", day_head, day_head)
            day_to = day_head

        war = 6
        out_nj = ""
        if self.include_jiangsu and self.include_nanjing:
            war = 4
            out_nj = "2"
        elif self.include_jiangsu and (not self.include_nanjing):
            war = 5
            out_nj = "1"

        time_interval = day_head + "---" + day_tail
        resp = ""
        if self.include_jiangsu:
            resp = requests.post(url=self.host + "/kyfx/info/jqkeyuanpaiming.jsp",
                                 params={"var": war, "data": time_interval, "outNj": out_nj},
                                 headers={"cookie": self.cookie})
        else:
            resp = requests.post(url=self.host + "/kyfx/info/jqkeyuanpaiming.jsp",
                                 params={"var": war, "data": time_interval},
                                 headers={"cookie": self.cookie})

        resp = requests.post(url=self.host + "/kyfx/info/jqkeyuanpaiming1.jsp",
                             params={"war": war, "outNj": out_nj}, headers={"cookie": self.cookie})
        resp.raise_for_status()

        resp = requests.post(url=self.host + "/kyfx/info/xiazai.jsp",
                             params={"var": "jqkeyuan"}, headers={"cookie": self.cookie})
        resp.raise_for_status()

        file_name = self.data_path + "\\" + day_head + "--" + day_tail + ".xls"
        file = open(file_name, "wb+")
        file.write(resp.content)
        file.close()

    def process_data(self, **kwargs):
        # combine data in [day_head, day_tail)
        if ("day_head" not in kwargs) or ("day_tail" not in kwargs):
            print("Arguments must contain: day_head, day_tail")
            return
        day_head = kwargs["day_head"]
        day_tail = kwargs["day_tail"]

        day_it = day_head
        cnt = 2
        day_idx = day_it.strftime("%Y-%m-%d")
        day_t = day_tail.strftime("%Y-%m-%d")

        wb = Workbook()
        ws = wb.active
        while day_idx != day_t:
            print("Processing...", day_idx)
            file_path = self.data_path + "\\" + day_idx + "--" + day_idx + ".xls"
            self.__combine_data(file_path, cnt, day_idx, ws)
            day_it = day_it + datetime.timedelta(days=1)
            cnt += 1
            day_idx = day_it.strftime("%Y-%m-%d")
        jiangsu = "剔除江苏"
        nanjing = "剔除南京"
        if self.include_nanjing:
            nanjing = "包含南京"
        if self.include_jiangsu:
            jiangsu = "包含江苏"

        t_h = day_head.strftime("%Y-%m-%d")
        t_t = (day_tail - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        wb.save(self.data_path + "\\" + "全市景区客源排名" + jiangsu + nanjing + t_h + "--" + t_t + ".xlsx")

    def __combine_data(self, file_path, cnt, day_idx, ws):
        book = ExcelTools.open_xls_as_xlsx(file_path=file_path)
        sheet = book.active
        ws.cell(row=cnt, column=1).value = day_idx
        if sheet.max_column == 1 or sheet.max_row ==1:
            return
        data = {}
        for r in range(2, sheet.max_row + 1):
            province = ExcelTools.erase_space(sheet.cell(row=r, column=1).value)
            if (not self.include_jiangsu) and (province == "江苏"):
                continue
            data[province] = sheet.cell(row=r,column=2).value

        c = 2
        for p in self.provinces:
            if (not self.include_jiangsu) and (p == "江苏"):
                continue
            ws.cell(row=1,column=c).value = p
            ws.cell(row=cnt,column=c).value = data.get(p)
            c = c + 1

    def pipeline(self, continuous=False, set_cookie=True, **kwargs):
        if ("year_head" not in kwargs) or ("year_tail" not in kwargs) or ("month_head" not in kwargs) \
                or ("month_tail" not in kwargs) or ("day_head" not in kwargs) or ("day_tail" not in kwargs) \
                or ("include_nanjing" not in kwargs) or ("include_jiangsu" not in kwargs):
            print("Arguments must contain: year_head, year_tail, month_head, month_tail, day_head, day_tail")
            return
        if set_cookie:
            self.set_cookie()
        else:
            if not self.login():
                return

        year_head = kwargs["year_head"]
        year_tail = kwargs["year_tail"]
        month_head = kwargs["month_head"]
        month_tail = kwargs["month_tail"]
        day_head = kwargs["day_head"]
        day_tail = kwargs["day_tail"]
        self.include_nanjing = kwargs["include_nanjing"]
        self.include_jiangsu = kwargs["include_jiangsu"]
        day_h = datetime.date(year_head, month_head, day_head)
        day_t = datetime.date(year_tail, month_tail, day_tail)
        if continuous:
            day_h_str = day_h.strftime("%Y-%m-%d")
            day_t_str = day_t.strftime("%Y-%m-%d")
            print("Downloading..." + day_h_str + " to " + day_t_str)
            self.fetch_data(day_head=day_h_str, day_tail=day_t_str)
        else:
            day_t = day_t + datetime.timedelta(days=1)
            day_h_str = day_h.strftime("%Y-%m-%d")
            day_t_str = day_t.strftime("%Y-%m-%d")
            while day_h_str != day_t_str:
                self.fetch_data(day_head=day_h_str, day_tail=day_h_str)
                print("Downloading...", day_h_str)
                day_h = day_h + datetime.timedelta(days=1)
                day_h_str = day_h.strftime("%Y-%m-%d")
            self.process_data(day_head=datetime.date(year_head, month_head, day_head), day_tail=day_t)


if __name__ == "__main__":
    downloader = QuanShiJingQuKeYuanPaiMing()
    downloader.pipeline(continuous=False, set_cookie=True, year_head=2016, year_tail=2016, month_head=9, month_tail=9,
                        day_head=1, day_tail=4, include_jiangsu=False, include_nanjing=False)